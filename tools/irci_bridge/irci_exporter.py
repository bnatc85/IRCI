#!/usr/bin/env python3
from pathlib import Path
import argparse, json, os
import pandas as pd
from datetime import datetime, timezone
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, numbers
from openpyxl.formatting.rule import FormulaRule
import numpy as np

CANONICAL_SHEETS = ["FILINGS","PRICES","FACTORS","EVENTS","NEWS"]

def latest_csv(root: Path, pattern: str) -> Path | None:
    files = sorted(root.glob(pattern))
    return files[-1] if files else None

def load_csvs_from_dir(d: Path):
    frames = {}
    for name in CANONICAL_SHEETS:
        p = d / f"{name}.csv"
        frames[name] = pd.read_csv(p) if p.exists() else pd.DataFrame()
    return frames

def coerce_dates(frames: dict):
    for nm in ["FILINGS","PRICES","FACTORS","EVENTS","NEWS"]:
        if not frames.get(nm, pd.DataFrame()).empty:
            for col in [c for c in frames[nm].columns if "Date" in c or c in ("WindowStart","WindowEnd")]:
                try:
                    frames[nm][col] = pd.to_datetime(frames[nm][col]).dt.date
                except Exception:
                    pass

def build_playbook(dials: pd.DataFrame,
                   comp_pipeline: pd.DataFrame,
                   comp_weighted: pd.DataFrame,
                   valuation: pd.DataFrame | None,
                   target_closure: float) -> pd.DataFrame:
    """
    Decision matrix that:
      - shows dials + both composite flavors
      - converts valuation_gap_pct into Valuation_Opportunity_$
      - flags the lowest dial and suggests a play
    """
    # Normalize column names
    d = dials.copy()
    # Expected DIALS columns: Quarter, Ticker, Coverage_pct, Trust_pct, Liquidity_pct, Valuation_pct
    # Add both composite flavors (joined on Quarter+Ticker)
    p = comp_pipeline.rename(columns={"CompositePct":"Composite_pipeline"})[["Quarter","Ticker","Composite_pipeline"]]
    w = comp_weighted.rename(columns={"CompositePct":"Composite_weighted"})[["Quarter","Ticker","Composite_weighted"]]
    out = d.merge(p, on=["Quarter","Ticker"], how="left").merge(w, on=["Quarter","Ticker"], how="left")

    # If we have valuation with EV + valuation_gap_pct, compute $ opportunity
    if valuation is not None and not valuation.empty:
        v = valuation.copy()

        # Flexible parsing: accept 17.2 or "17.2%" or 0.172; normalize to fraction
        if "valuation_gap_pct" in v.columns:
            g = pd.to_numeric(v["valuation_gap_pct"].astype(str).str.replace("%","", regex=False), errors="coerce")
            # If median absolute value > 1, it’s in percent points (e.g., 17.2); convert to fraction
            if g.abs().median(skipna=True) > 1.0:
                g = g / 100.0
            v["valuation_gap_pct"] = g

        # Label quarter to join on DIALS
        if "quarter_end" in v.columns:
            v["Quarter"] = pd.to_datetime(v["quarter_end"], utc=True, errors="coerce").dt.to_period("Q-DEC").astype(str)

        keep_cols = ["Quarter","ticker","enterprise_value","valuation_gap_pct"]
        v2 = v[[c for c in keep_cols if c in v.columns]].rename(columns={"ticker":"Ticker"})
        out = out.merge(v2, on=["Quarter","Ticker"], how="left")

        # Opportunity: only when undervalued (gap < 0). Compression risk when overvalued (gap > 0).
        out["enterprise_value"]     = pd.to_numeric(out["enterprise_value"], errors="coerce")
        out["valuation_gap_pct"]    = pd.to_numeric(out["valuation_gap_pct"], errors="coerce")

        tgt = float(target_closure)  # use the CLI flag, not a hard-coded 0.50

        # gap is a fraction (e.g., -0.112 means 11.2% below peer)
        out["Valuation_Opportunity_$"] = out["enterprise_value"] * (-out["valuation_gap_pct"].clip(upper=0)) * tgt
        out["Compression_Risk_$"]      = out["enterprise_value"] * ( out["valuation_gap_pct"].clip(lower=0)) * tgt

        # Optional: express opportunity as % of EV and zero-out tiny (<1%EV) noise
        out["Opportunity_%EV"] = 100.0 * (out["Valuation_Opportunity_$"] / out["enterprise_value"])
    else:
        out["Valuation_Opportunity_$"] = pd.NA
        out["Compression_Risk_$"] = pd.NA
        out["Opportunity_%EV"] = pd.NA


    # Lowest dial + simple playbook
    dial_cols = ["Coverage_pct","Trust_pct","Liquidity_pct","Valuation_pct"]
    out["LowestDial"] = out[dial_cols].idxmin(axis=1)
    def play_for(dial: str) -> str:
        if dial == "Coverage_pct":
            return "Coverage: 8-K hygiene, briefings, calendar discipline"
        if dial == "Trust_pct":
            return "Trust: proactive calm events, top-tier media, clarity"
        if dial == "Liquidity_pct":
            return "Liquidity: market structure, MM engagement, split eval"
        if dial == "Valuation_pct":
            return "Valuation: narrative + guidance; targets vs peers"
        return "Balanced: maintain cadence"
    out["Play"] = out["LowestDial"].map(play_for)

    # Cosmetic ordering
    cols = [
        "Quarter","Ticker",
        "Coverage_pct","Trust_pct","Liquidity_pct","Valuation_pct",
        "Composite_pipeline","Composite_weighted",
        "Valuation_Opportunity_$","Compression_Risk_$","Opportunity_%EV",
        "LowestDial","Play"
    ]
    return out[cols]

def write_excel_all(excel_path: Path,
                    frames: dict,
                    dials: pd.DataFrame,
                    comp_p: pd.DataFrame,
                    comp_w: pd.DataFrame,
                    playbook: pd.DataFrame):
    # One writer (openpyxl) for everything
    with pd.ExcelWriter(excel_path, engine="openpyxl", mode="w") as xw:
        # Canonical sheets first (so templates expecting them still work)
        for name, df in frames.items():
            df.to_excel(xw, sheet_name=name, index=False)
        # Our analysis sheets
        dials.to_excel(xw, sheet_name="DIALS", index=False)
        comp_p.to_excel(xw, sheet_name="COMPOSITE_pipeline", index=False)
        comp_w.to_excel(xw, sheet_name="COMPOSITE_weighted", index=False)
        playbook.to_excel(xw, sheet_name="PLAYBOOK", index=False)

    # Post-formatting: highlight lowest dial in DIALS, currency in PLAYBOOK
    wb = load_workbook(excel_path)
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # DIALS: mark the minimum per row across C:F (Coverage..Valuation)
    ws = wb["DIALS"]
    max_row = ws.max_row
    # Apply row-wise rule C2:F{n}: cell = MIN($C{r}:$F{r})
    for r in range(2, max_row+1):
        for col_letter in ("C","D","E","F"):
            cell_range = f"{col_letter}{r}"
            formula = f'{col_letter}{r}=MIN($C{r}:$F{r})'
            ws.conditional_formatting.add(cell_range, FormulaRule(formula=[formula], fill=red))

    # PLAYBOOK: currency format for both $ columns and % for Opportunity_%EV
    if "PLAYBOOK" in wb.sheetnames:
        ws2 = wb["PLAYBOOK"]
        headers = {ws2.cell(row=1, column=c).value: c for c in range(1, ws2.max_column+1)}

        for hdr in ("Valuation_Opportunity_$", "Compression_Risk_$"):
            if hdr in headers:
                cidx = headers[hdr]
                for r in range(2, ws2.max_row+1):
                    ws2.cell(row=r, column=cidx).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

        if "Opportunity_%EV" in headers:
            pidx = headers["Opportunity_%EV"]
            for r in range(2, ws2.max_row+1):
                ws2.cell(row=r, column=pidx).number_format = numbers.FORMAT_PERCENTAGE_00


    wb.save(excel_path)

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--from-exports", required=True,
                    help="Dir with canonical CSVs: FILINGS/PRICES/FACTORS/EVENTS/NEWS/DIALS/COMPOSITE.csv")
    ap.add_argument("--outputs-dir", default="./outputs",
                    help="Dir where per-step outputs like valuation*.csv live (for gap & EV)")
    ap.add_argument("--excel", required=True,
                    help="Path to Excel to write (e.g., artifacts/IRCI_lowcode_template.xlsx)")
    ap.add_argument("--manifest", default="manifest.json")
    ap.add_argument("--gap-closure-target", type=float, default=0.50,
                    help="Target proportion of valuation gap to close (0–1), default 0.50")
    args = ap.parse_args()

    exports_dir = Path(args.from_exports)
    outputs_dir = Path(args.outputs_dir)
    excel_path  = Path(args.excel)

    # Canonical frames (pass-through tables)
    frames = load_csvs_from_dir(exports_dir)
    coerce_dates(frames)

    # DIALS + both composites (we expect both CSVs to exist now)
    dials  = pd.read_csv(exports_dir / "DIALS.csv")
    comp_p = pd.read_csv(exports_dir / "COMPOSITE.csv")              # pipeline
    comp_w = pd.read_csv(exports_dir / "COMPOSITE_WEIGHTED.csv")     # weighted

    # Pull latest valuation to get EV + valuation_gap_pct (optional)
    v_path = latest_csv(outputs_dir, "valuation*.csv")
    valuation = pd.read_csv(v_path) if v_path else pd.DataFrame()

    playbook = build_playbook(dials, comp_p, comp_w, valuation, args.gap_closure_target)

    # Write everything to one workbook and format
    write_excel_all(excel_path, frames, dials, comp_p, comp_w, playbook)

    manifest = {
        "export_time_utc": datetime.now(timezone.utc).isoformat(),
        "source_dir": os.path.abspath(args.from_exports),
        "targets": {"excel": str(excel_path)},
        "tables_written": {k: int(len(v)) for k,v in frames.items()},
        "extras": {
            "has_weighted": True,
            "gap_closure_target": args.gap_closure_target,
            "valuation_csv": str(v_path) if v_path else None
        }
    }
    with open(args.manifest, "w") as f:
        json.dump(manifest, f, indent=2)
    print(f"Wrote: {excel_path}")
    print("Manifest:", os.path.abspath(args.manifest))

if __name__ == "__main__":
    main()
